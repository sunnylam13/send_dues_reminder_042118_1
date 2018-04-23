# -*- coding: utf-8 -*-

#! python3

# USAGE
# python3 send_dues_reminder_042118_1.py -e "EMAIL" -p "PASSWORD"
# python3 send_dues_reminder_042118_1.py -e "avatar.sage7@gmail.com" -p "PASSWORD"

import logging
logging.basicConfig(level=logging.DEBUG, format=" %(asctime)s - %(levelname)s - %(message)s")
# logging.disable(logging.CRITICAL)

import openpyxl,smtplib
import optparse
import pprint

def unpaid_members_analysis_1():
	# currently just using the global variables
	
	for r in range( 2,sheet.max_row + 1 ):
		payment = sheet.cell( row=r, column=lastCol ).value
		logging.debug( 'payment status is:  %s' % str(payment) )

		if payment != 'paid':
			name = sheet.cell(row=r,column=1).value
			logging.debug( 'name is:  %s' % str(name) )
			email = sheet.cell(row=r,column=2).value
			logging.debug( 'email is:  %s' % str(email) )
			unpaidMembers[name] = email
			logging.debug( 'email pushed into unpaidMembers dict' )

	logging.debug( 'unpaidMembers dict is:  ' )
	logging.debug( unpaidMembers )

def email_login_1(email,pwd):
	try:
		# smtpObj = smtplib.SMTP('smtp.gmail.com',587) # open smtp connection
		smtpObj = smtplib.SMTP_SSL('smtp.gmail.com',465) # open smtp connection
		smtpObj.ehlo()
		logging.debug( 'ehlo() worked.' )
		# smtpObj.starttls() # disable if using SMTP_SSL()
		# logging.debug( 'starttls() worked.' )

		# logging.debug( 'pwd was:  %s' % str(sys.argv[1]) )
		smtpObj.login( str(email), str(pwd) )
		logging.debug( 'Email provider login successful.' )
	except Exception as e:
		logging.debug( "There was an exception:  %s" % str(e) )
		raise e

	# optional test
	# logging.debug( "Listing folders:  " )
	# logging.debug( pprint.pprint(smtpObj.list_folders()) )

	return smtpObj # for use outside of function

# send out reminder emails

def send_reminder_emails_1(dict,smtpObj,email):
	for name,email in unpaidMembers.items():
		body = """
			Subject: %s dues unpaid.\n
			Dear %s,\n
			Records show that you have not paid dues for %s.\n
			Please make this payment as soon as possible.\n
			Thank you!
			""" % (latestMonth,name,latestMonth)
		print( 'Sending email to %s...' % str(email) )
		sendmailStatus = smtpObj.sendmail( str(email),email,body )

		if sendmailStatus != {}: # if the send mail status has failed send targets, which means a non-empty dict is returned
			print( "There was a problem sending email to %s:  %s" % (email,sendmailStatus) )

	smtpObj.quit() # close smtp connection when done

def main():

	parser = optparse.OptionParser( 'usage %prog ' + '-e <email> ' + '-p <password>' ) # make sure there's a space at the end of each string except the last one 
	parser.add_option('-e', dest='email', type='string', help='specify login email')
	parser.add_option('-p', dest='pwd', type='string', help='specify email service provider password')
	(options, args) = parser.parse_args()

	if options.email == None:
		print (parser.usage)
		exit(0)
	else:
		email_login = options.email
		logging.debug( 'Proxy is:  %s' % str(url_target) )

	if options.pwd == None:
		print (parser.usage)
		exit(0)
	else:
		email_pwd = options.pwd
		logging.debug( 'Proxy is:  %s' % str(url_target) )

	# open the sheet and get latest due status

	wb = openpyxl.load_workbook('duesRecords.xlsx') # normally I'd code this to take a command line input
	logging.debug( 'Workbook opened' )

	sheet = wb['Sheet1']
	logging.debug( 'Sheet title is:  %s' % str(sheet.title) )

	lastCol = sheet.max_column
	logging.debug( 'Last column is:  %s' % str(lastCol) )

	latestMonth = sheet.cell(row=1,column=lastCol).value
	logging.debug( 'Latest month is:  %s' % str(latestMonth) )

	unpaidMembers = {}

	# check each member's payment status	
	unpaid_members_analysis_1()

	# log into email account
	smtpObj = email_login_1(email_login,email_pwd)

	# send out reminder emails
	send_reminder_emails_1(unpaidMembers,smtpObj,email_login)

if __name__ == '__main__':
	main()
